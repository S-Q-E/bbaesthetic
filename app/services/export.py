from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.repositories import ApplicationRepository


@dataclass(slots=True, frozen=True)
class ExportResult:
    path: Path
    filename: str


class ExportService:
    HEADERS = (
        "ID",
        "Имя",
        "Возраст",
        "Город",
        "Цель похудения",
        "Телефон",
        "Username",
        "Telegram ID",
        "Дата",
    )

    def __init__(self, repository: ApplicationRepository, tmp_dir: Path) -> None:
        self.repository = repository
        self.tmp_dir = tmp_dir

    async def export_applications(self) -> ExportResult | None:
        applications = await self.repository.get_all()
        if not applications:
            return None

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Applications"
        worksheet.freeze_panes = "A2"

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        centered_alignment = Alignment(vertical="center", horizontal="center")

        worksheet.append(self.HEADERS)

        for index, header in enumerate(self.HEADERS, start=1):
            cell = worksheet.cell(row=1, column=index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered_alignment

        for application in applications:
            worksheet.append(
                [
                    application.id,
                    application.full_name,
                    application.age,
                    application.city,
                    f"-{application.target_weight_loss} кг",
                    application.phone,
                    f"@{application.telegram_username}"
                    if application.telegram_username
                    else "",
                    application.telegram_id,
                    application.created_at.strftime("%Y-%m-%d %H:%M"),
                ]
            )

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value_length = len(str(cell.value or ""))
                if value_length > max_length:
                    max_length = value_length
                cell.alignment = Alignment(vertical="center")
            worksheet.column_dimensions[column_letter].width = max(max_length + 2, 14)

        filename = f"applications_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
        filepath = self.tmp_dir / f"tmp_{filename}"
        workbook.save(filepath)

        return ExportResult(path=filepath, filename=filename)
